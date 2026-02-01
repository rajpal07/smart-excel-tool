import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import re
import io

class ExcelHandler:
    def __init__(self):
        pass

    def get_sheet_info(self, file):
        """
        Returns a dictionary of {sheet_name: row_count} (approx) and list of sheet names.
        """
        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            return sheet_names
        except Exception as e:
            print(f"Error reading workbook: {e}")
            return []

    def get_cell_hex(self, cell):
        """
        Robustly extracts the hex color background of a cell.
        Handles RGB, Theme (approximate), and Index.
        Returns None if no fill.
        """
        if not cell.fill: return None
        # We accept any patternType if start_color is present
        color = cell.fill.start_color
        if not color: return None
        
        # 1. RGB
        if color.type == 'rgb':
            if len(color.rgb) == 8:
                alpha = color.rgb[:2]
                rgb = color.rgb[2:]
                # Check for Transparent Black (00000000) -> No Color
                if alpha == '00' and rgb == '000000': return None 
                
                # Check for White (00FFFFFF or FFFFFFFF or just FFFFFF)
                # If it's pure white, treat as "No Color" so we can overwrite it.
                if rgb.upper() == 'FFFFFF': return None
                
                # If alpha is 00 but rgb has color (e.g. 00FF0000), it's likely a quirk or actually colored.
                # Treat as colored to be safe (preserve it).
                return rgb
            return color.rgb
            
        # 2. Theme or Indexed
        if color.type in ['theme', 'indexed']:
            # We can't easily resolve theme colors without workbook context
            # But usually we don't want to override theme colors either?
            # Or return a placeholder effectively saying "Has Color"
            return "THEME" 
            
        return None

    def _get_trimmed_bounds(self, ws):
        """
        Finds the true last row and column with data (ignoring empty styling).
        Returns (max_row, max_col)
        """
        max_row = ws.max_row
        max_col = ws.max_column
        
        # 1. Find last row with data
        real_max_row = 1
        # Scan blocks of rows backwards for speed? Or just iterate.
        # Check from bottom up
        for r in range(max_row, 0, -1):
            # Check row r
            row_has_data = False
            for c in range(1, max_col + 1):
                cell_val = ws.cell(row=r, column=c).value
                if cell_val is not None and str(cell_val).strip() != "":
                    row_has_data = True
                    break
            if row_has_data:
                real_max_row = r
                break
                
        # 2. Find last col with data
        real_max_col = 1
        for c in range(max_col, 0, -1):
            col_has_data = False
            for r in range(1, real_max_row + 1):
                cell_val = ws.cell(row=r, column=c).value
                if cell_val is not None and str(cell_val).strip() != "":
                    col_has_data = True
                    break
            if col_has_data:
                real_max_col = c
                break
                
        return real_max_row, real_max_col

    def read_sheet_with_styles(self, file, sheet_name, limit=None):
        """
        Reads the sheet, extracting values AND background colors.
        Trims empty rows/cols to improve performance.
        limit=None means read all data rows.
        Returns a list of rows, where each cell is {'value': val, 'bg_color': hex}.
        """
        try:
            wb = openpyxl.load_workbook(file, data_only=True) # data_only=True to get values not formulas
            if sheet_name not in wb:
                return None
            
            ws = wb[sheet_name]
            
            # Calculate trimmed bounds
            max_r, max_c = self._get_trimmed_bounds(ws)
            
            final_max_row = max_r
            if limit:
                final_max_row = min(limit, max_r)
            
            rows_data = []
            
            # Iter_rows with max_col ensures we don't fetch 16000 empty columns
            for row in ws.iter_rows(min_row=1, max_row=final_max_row, max_col=max_c):
                row_data = []
                for cell in row:
                    val = cell.value
                    # FORCE STRING to avoid PyArrow errors
                    val_str = str(val) if val is not None else None
                    
                    # Use helper
                    hex_c = self.get_cell_hex(cell)
                    bg_color = f"#{hex_c}" if hex_c else None
                    
                    row_data.append({
                        'value': val_str,
                        'bg_color': bg_color
                    })
                rows_data.append(row_data)
            return rows_data
        except Exception as e:
            print(f"Error reading sheet styles: {e}")
            return []

    def detect_threshold_rows(self, rows_data):
        """
        Analyzes the top extracted rows to find the 'Threshold Block'.
        Heuristics:
        - Consecutive rows.
        - Have background colors.
        - Often near the top (rows 0-10).
        - Contain numbers or ranges.
        
        Returns a list of row indices.
        """
        colored_rows = []
        for idx, row in enumerate(rows_data):
            # Check if row has significant coloring (e.g., > 10% cells colored? Or ANY cell colored?)
            # User said "colourful rows above the data".
            # Usually specific columns have color.
            # Let's count colored cells with content.
            colored_cells = [c for c in row if c['bg_color'] and c['value'] is not None]
            if len(colored_cells) > 0:
                colored_rows.append(idx)
        
        # Group typical blocks? 
        # For now, just return all rows that have *some* coloring in the top section.
        return colored_rows

    def parse_threshold_rules(self, rows_data, row_indices, col_idx):
        """
        Extracts rules for a specific column from the identified threshold rows.
        Returns list of rules: {'min': float, 'max': float, 'color': hex}
        """
        rules = []
        
        for r_idx in row_indices:
            if r_idx >= len(rows_data): continue
            row = rows_data[r_idx]
            if col_idx >= len(row): continue
            
            cell = row[col_idx]
            val = cell['value']
            color = cell['bg_color']
            
            if not color or val is None:
                continue
                
            # Parse contents
            val_str = str(val).strip()
            
            # Case 1: Range "5-7"
            # Regex for "number - number"
            range_match = re.match(r'([\d\.]+)\s*-\s*([\d\.]+)', val_str)
            if range_match:
                try:
                    min_val = float(range_match.group(1))
                    max_val = float(range_match.group(2))
                    rules.append({'min': min_val, 'max': max_val, 'color': color, 'priority': 2}) # Range = High priority
                    continue
                except:
                    pass
            
            # Case 2: Single Number "8.5" (Usually Min or Max?)
            # Context implies: "Red - 8.5" -> > 8.5? or < 8.5?
            # User said: "colour the cells which are above the respective threshold value"
            # So single value = MINIMUM threshold.
            try:
                # Handle operators if present
                clean_val = val_str.replace('>', '').replace('<', '').replace('=', '').strip()
                float_val = float(clean_val)
                
                if '<' in val_str:
                     # Explicit "Less than"
                     rules.append({'min': float('-inf'), 'max': float_val, 'color': color, 'priority': 2})
                else:
                     # Default > Value
                     rules.append({'min': float_val, 'max': float('inf'), 'color': color, 'priority': 1})
            except:
                continue
                
        # Sort by priority then by value?
        # Ranges (prio 2) should be checked first?
        # Or simple: Check STRICTEST first?
        # Actually user said "colour cells ABOVE threshold".
        # If we have Red > 8, Green > 3.
        # A value 9 is > 3 and > 8. Should be Red.
        # So we should sort by DESCENDING threshold value for single-min rules.
        
        rules.sort(key=lambda x: x.get('min', 0), reverse=True)
        return rules

    def clone_fill(self, fill):
        """Creates a deep copy of a PatternFill object."""
        if not fill: return None
        return PatternFill(
            fill_type=fill.fill_type,
            start_color=fill.start_color,
            end_color=fill.end_color,
            fgColor=fill.fgColor, 
            bgColor=fill.bgColor
        )

    def _is_white_color(self, hex_color):
        """
        Returns True if color is white or near-white.
        Used to determine if a cell should be considered "uncolored".
        """
        if not hex_color:
            return True
        # Remove # if present and normalize
        clean = hex_color.replace('#', '').upper()
        # Check common white variants
        # FFFFFF = white, 00000000 = no fill, FEFEFE = near-white
        return clean in ['FFFFFF', 'FFFFFFFF', 'FEFEFE', '00000000', 'FEFEFEFE']

    def parse_cell_value(self, val_str, col_name=""):
        """
        Parses a cell string to extract rule parameters.
        Returns dict: {'type': 'range'|'less'|'greater'|'implicit'|'ambiguous_less', 'min': float, 'max': float, 'original': str}
        """
        # 1. Clean comments: #, *, a, S, sat, ^
        # User requested: "in the # * list add S , sat ^"
        
        # Original clean: re.split(r'sat|[#\*aS\^]', val_str)[0].strip()
        # We also need to handle '%'. If we treat % as unit to ignore.
        # "80%" -> "80".
        # We could add % to the split list? Or just strip it.
        # Split logic: "80% S" -> split on % -> "80". Correct.
        # "5-10%" -> split on % -> "5-10". Correct.
        # "5%-10%" -> split on % -> "5". FAIL range.
        
        # Better strategy: Replace % with empty string first, THEN split on comments.
        # "80% S" -> "80 S" -> split S -> "80".
        # "5%-10%" -> "5-10" -> split S (none) -> "5-10".
        
        val_no_percent = val_str.replace('%', '').replace(',', '').replace('cfu/100 ml', '').replace('(in 100 ml) #9', '')
        clean_str = re.split(r'sat|[#\*aS\^]', val_no_percent)[0].strip()
        
        # 2. Check Range "6.5-8.5"
        range_match = re.match(r'^([\d\.]+)\s*-\s*([\d\.]+)$', clean_str)
        if range_match:
            try:
                return {
                    'type': 'range',
                    'min': float(range_match.group(1)),
                    'max': float(range_match.group(2))
                }
            except: pass
            
        # 3. Check Explicit < or >
        if '<' in clean_str:
            try:
                val = float(clean_str.replace('<', '').replace('=', '').strip())
                return {'type': 'ambiguous_less', 'val': val} 
            except: pass
            
        if '>' in clean_str:
            try:
                val = float(clean_str.replace('>', '').replace('=', '').strip())
                return {'type': 'greater', 'min': val, 'max': float('inf')}
            except: pass
            
        # 4. Implicit (Just a number) -> Assume > Value
        try:
            val = float(clean_str)
            return {'type': 'implicit', 'min': val, 'max': float('inf')}
        except:
            return None

    def resolve_column_rules(self, rules):
        """
        Resolves 'ambiguous_less' rules (< Val) based on context of other rules in column.
        Heuristic:
        - If a 'range' rule (min-max) exists:
            - If rule.val <= range.min: Treat as LESS THAN (Lower Bound).
            - Else: Treat as GREATER THAN (Upper Limit).
        - If NO range rule exists:
            - Assume 'Pollutant Model' (Exceedance). Treat as GREATER THAN.
        """
        # Find Range Context
        range_min = None
        for r in rules:
            if r['type'] == 'range':
                # Use the lowest range min found if multiple
                if range_min is None or r['min'] < range_min:
                    range_min = r['min']
        
        resolved_rules = []
        for r in rules:
            if r['type'] == 'ambiguous_less':
                val = r['val']
                # Decide Context
                is_lower_bound = False
                
                if range_min is not None:
                     # Range Context: Is it on the left side?
                     # e.g. Range 6.5-8.5. Val 5.5. 5.5 < 6.5 -> Lower Bound.
                     # e.g. Range 6.5-8.5. Val 9.0. 9.0 > 8.5 -> Upper Bound.
                     # Strict Compare? Use epsilon?
                     if val <= range_min + 0.000001:
                         is_lower_bound = True
                else:
                     # NO Range Context.
                     # Heuristic: 
                     # If val is "small" (e.g. < 1), likely Limit of Reporting (LOR). Treat as > Val.
                     # If val is "large" (e.g. > 5), likely a Lower Limit (e.g. Dissolved Oxygen < 80). Treat as < Val.
                     if val > 5.0:
                         is_lower_bound = True
                
                # Apply Decision
                if is_lower_bound:
                    # Treat as < Val
                    r['type'] = 'less'
                    r['min'] = float('-inf')
                    r['max'] = val
                else:
                    # Treat as > Val (Limit Exceedance)
                    r['type'] = 'greater'
                    r['min'] = val
                    r['max'] = float('inf')
            
            resolved_rules.append(r)
        return resolved_rules

    def apply_thresholds(self, file_path, sheet_name, target_columns_indices, threshold_rows_indices, output_path, data_start_row=None, header_row_idx=None, manual_overrides=None):
        """
        The Heavy Lifter.
        """
        try:
            # FORCE data_only=True to ensure we save VALUES.
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except:
             wb = openpyxl.load_workbook(file_path, data_only=True)
             
        if sheet_name not in wb: return None
        ws = wb[sheet_name]
        
        # Get Column Names if header row provided (for context logic)
        col_names = {}
        if header_row_idx is not None:
             # openpyxl uses 1-based indexing for args, but we have 0-based idx?
             # header_row_idx is likely 0-based from pandas read.
             # So row = header_row_idx + 1
             for col_idx in target_columns_indices:
                 cell = ws.cell(row=header_row_idx+1, column=col_idx+1)
                 col_names[col_idx] = str(cell.value) if cell.value else ""
        
        # 1. Determine Data Start Row
        if data_start_row is None:
             max_thresh = max(threshold_rows_indices) if threshold_rows_indices else 0
             data_start_row = max_thresh + 2
             
             # Attempt Heuristic: Look for "Assessment Guidelines" or similar
             found_guideline = False
             for i, row in enumerate(ws.iter_rows(max_row=50, values_only=True)):
                 row_str = " ".join([str(c) for c in row if c]).lower()
                 if "assessment guideline" in row_str or "guideline" in row_str:
                     data_start_row = i + 2 # Start *after* this row (1-based + 1)
                     found_guideline = True
                     print(f"Detected Data Start Row after 'Assessment Guidelines' at: {data_start_row}")
                     break
             
             if not found_guideline:
                 print(f"No guideline row found, starting at {data_start_row} based on selection.")
        
        # Calculate trimmed bounds for processing
        max_r, max_c = self._get_trimmed_bounds(ws)
        
        # Parse Rules
        col_rules = {}
        # Ensure rows are sorted by index
        sorted_rows_indices = sorted(threshold_rows_indices)
        
        for col_idx in target_columns_indices:
            rules = []
            c_name = col_names.get(col_idx, "")
            
            for r_idx in sorted_rows_indices:
                cell = ws.cell(row=r_idx+1, column=col_idx+1) 
                
                if not self.get_cell_hex(cell): continue
                
                fill_copy = self.clone_fill(cell.fill)
                val = cell.value
                if val is None: continue
                val_str = str(val).strip()
                
                parsed = self.parse_cell_value(val_str, col_name=c_name)
                if parsed:
                    # Append in order of rows
                    rule = parsed
                    rule['fill'] = fill_copy
                    rule['row_idx'] = r_idx # Debug info
                    rules.append(rule)
            
            # Resolve Ambiguities (e.g. < Val) based on gathered context
            rules = self.resolve_column_rules(rules)
            
            # DO NOT SORT by value/priority. Respect Row Order.
            col_rules[col_idx] = rules
            
        # 2. Apply to Data
        count = 0
        
        # Use trimmed bounds for iteration
        iterator = ws.iter_rows(min_row=data_start_row, max_row=max_r, max_col=max_c)
        
        for row in iterator:
            for col_idx in target_columns_indices:
                if col_idx >= len(row): continue
                cell = row[col_idx]
                
                # --- Manual Overrides ---
                # manual_overrides is dict: {(row_idx_0_based, col_idx_0_based): 'HEX' or None}
                current_row_idx = cell.row - 1 # 0-based
                
                if manual_overrides and (current_row_idx, col_idx) in manual_overrides:
                    override_hex = manual_overrides[(current_row_idx, col_idx)]
                    if override_hex:
                         # Apply Override
                         cell.fill = PatternFill(start_color="FF" + override_hex, end_color="FF" + override_hex, fill_type="solid")
                         count += 1
                         continue
                    else:
                         # Explicitly Clear Color (if override is None/'')
                         cell.fill = PatternFill(fill_type=None)
                         continue
                # -----------------------------


                # --- Preserve Existing Non-White Color ---
                # If the cell already has a non-white color, skip threshold application
                existing_color = self.get_cell_hex(cell)
                if existing_color and not self._is_white_color(existing_color):
                    continue
                # ------------------------------------


                val = cell.value
                
                try:
                    if isinstance(val, str):
                        # Handle < LOR logic:
                        # User wants <0.01 to be TREATED as less than 0.01.
                        # If we just strip <, it becomes 0.01, which is >= 0.01 (True).
                        # So we subtract a tiny epsilon.
                        if '<' in val:
                             f_val = float(val.replace('<','').replace('>','').replace(',','')) - 1e-9
                        else:
                             f_val = float(val.replace('<','').replace('>','').replace(',',''))
                    else:
                        f_val = float(val)
                except:
                    continue 
                    
                if col_idx in col_rules:
                    # Strictest Rule Wins Logic:
                    # 1. Find all matching rules.
                    matches = []
                    for rule in col_rules[col_idx]:
                        is_range = (rule.get('type') == 'range')
                        min_ok = (f_val >= rule['min']) if is_range else (f_val > rule['min'])
                        max_ok = (f_val <= rule['max']) if is_range else (f_val < rule['max'])
                        
                        if min_ok and max_ok:
                            matches.append(rule)
                    
                    if matches:
                        # 2. Sort by Strictness
                        # Key: (Strictness Score, -RowIndex)
                        # Strictness Score:
                        #   For > / Range: Higher min is stricter. Score = rule['min']
                        #   For < : Lower max is stricter. Score = -rule['max']
                        # Row Index:
                        #   Tie-breaker: Top-most row (smallest index) wins.
                        #   We want max() to pick it, so use -row_idx (e.g. -17 > -19).
                        
                        def strictness_key(r):
                            score = r['min']
                            # Heuristic for 'less' type: check if min is -inf
                            if r['min'] == float('-inf'):
                                score = -r['max']
                            return (score, -r.get('row_idx', 9999))
                            
                        # Sort and pick last (highest score/priority)
                        matches.sort(key=strictness_key)
                        winner = matches[-1]
                        
                        cell.fill = winner['fill']
                        count += 1
        
        print(f"Applied colors to {count} cells.")
        wb.save(output_path)
        return output_path

    def preview_thresholds(self, file_path, sheet_name, target_columns_indices, threshold_rows_indices, data_start_row=None, limit=None, header_row_idx=None, manual_overrides=None):
        """
        Returns a styled DataFrame showing how the rules WOULD apply to the first N rows of data.
        """
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]
        
        # Get Column Names
        col_names = {}
        if header_row_idx is not None:
             for col_idx in target_columns_indices:
                 cell = ws.cell(row=header_row_idx+1, column=col_idx+1)
                 col_names[col_idx] = str(cell.value) if cell.value else ""
        
        # Calculate trimmed bounds
        max_r, max_c = self._get_trimmed_bounds(ws)
        
        # 1. Determine Data Start (Dup logic)
        final_start_row = data_start_row
        if final_start_row is None:
             max_thresh = max(threshold_rows_indices) if threshold_rows_indices else 0
             final_start_row = max_thresh + 2
             for i, row in enumerate(ws.iter_rows(max_row=50, values_only=True)):
                 row_str = " ".join([str(c) for c in row if c]).lower()
                 if "assessment guideline" in row_str or "guideline" in row_str:
                     final_start_row = i + 2
                     break
        
        # Ensure start row is within bounds
        if final_start_row > max_r:
            return [], final_start_row

        # 2. Extract Rules (Same parsing logic)
        col_rules = {} 
        sorted_rows_indices = sorted(threshold_rows_indices)
        
        for col_idx in target_columns_indices:
            rules = []
            c_name = col_names.get(col_idx, "")
            
            for r_idx in sorted_rows_indices:
                cell = ws.cell(row=r_idx+1, column=col_idx+1)
                
                # Robust Hex
                hex_c = self.get_cell_hex(cell)
                if not hex_c: continue
                hex_color = hex_c
                
                val = cell.value
                if val is None: continue
                val_str = str(val).strip()
                
                parsed = self.parse_cell_value(val_str, col_name=c_name)
                if parsed:
                    rule = parsed
                    rule['hex'] = hex_color
                    rule['row_idx'] = r_idx # Critical for tie-breaking
                    rules.append(rule)
            
            # Resolve Ambiguities
            rules = self.resolve_column_rules(rules)
            
            # No Sort.
            col_rules[col_idx] = rules
            
        # 3. Read Data Rows for Preview
        preview_rows = []
        
        # Handle Limit
        final_max_row_limit = max_r
        if limit:
             final_max_row_limit = min(final_start_row + limit, max_r)
        
        # Use max_col to optimize
        iterator = ws.iter_rows(min_row=final_start_row, max_row=final_max_row_limit, max_col=max_c)
             
        consecutive_empty = 0
        MAX_EMPTY = 10
        
        for row in iterator:
            # Check for Empty Row (Optimization)
            is_empty = True
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != "":
                    is_empty = False
                    break
            
            if is_empty:
                consecutive_empty += 1
                if consecutive_empty >= MAX_EMPTY:
                    break
            else:
                consecutive_empty = 0 # Reset count if data found

            row_data = {}
            for i, cell in enumerate(row):
                val = cell.value
                row_data[f"Col {i}"] = {'value': str(val) if val is not None else "", 'bg': None}
                
                # --- Manual Overrides for Preview ---
                current_row_idx = cell.row - 1
                if manual_overrides and (current_row_idx, i) in manual_overrides:
                    override_hex = manual_overrides[(current_row_idx, i)]
                    row_data[f"Col {i}"]['bg'] = override_hex
                    continue
                
                # --- Check for Existing Color First ---
                existing_hex = self.get_cell_hex(cell)
                if existing_hex:
                    row_data[f"Col {i}"]['bg'] = existing_hex
                    continue

                if i in target_columns_indices:
                    try:
                        v_str = str(val if val is not None else "")
                        if '<' in v_str:
                            f_val = float(v_str.replace('<','').replace('>','').replace(',','')) - 1e-9
                        else:
                            f_val = float(v_str.replace('<','').replace('>','').replace(',',''))
                        
                        # Strictest Rule Logic for Preview
                        matches = []
                        for rule in col_rules[i]:
                            is_range = (rule.get('type') == 'range')
                            min_ok = (f_val >= rule['min']) if is_range else (f_val > rule['min'])
                            max_ok = (f_val <= rule['max']) if is_range else (f_val < rule['max'])
                            
                            if min_ok and max_ok:
                                matches.append(rule)
                                
                        if matches:
                            def strictness_key(r):
                                score = r['min']
                                if r['min'] == float('-inf'):
                                    score = -r['max']
                                return (score, -r.get('row_idx', 9999))
                                
                            matches.sort(key=strictness_key)
                            winner = matches[-1]
                            row_data[f"Col {i}"]['bg'] = winner['hex']
                    except: pass
            preview_rows.append(row_data)
            
        return preview_rows, final_start_row

    def preview_thresholds_from_data(self, rows_data, target_columns_indices, threshold_rows_indices, data_start_row=None, limit=None, header_row_idx=None, manual_overrides=None):
        """
        Generates preview using ALREADY LOADED rows_data.
        Avoids re-opening the file.
        """
        if not rows_data: return [], 0
        
        # Get Column Names
        col_names = {}
        if header_row_idx is not None and header_row_idx < len(rows_data):
             row_data = rows_data[header_row_idx]
             for col_idx in target_columns_indices:
                 if col_idx < len(row_data):
                     val = row_data[col_idx]['value']
                     col_names[col_idx] = str(val) if val else ""
        
        # 1. Determine Data Start (Dup logic, adapted for list)
        final_start_row = data_start_row
        if final_start_row is None:
             max_thresh = max(threshold_rows_indices) if threshold_rows_indices else 0
             final_start_row = max_thresh + 2
             
             # Scan top 50
             for i, row in enumerate(rows_data[:50]):
                 row_vals = [c['value'] for c in row if c['value']]
                 row_str = " ".join([str(v) for v in row_vals]).lower()
                 if "assessment guideline" in row_str or "guideline" in row_str:
                     final_start_row = i + 2
                     break
        
        if final_start_row > len(rows_data):
             final_start_row = len(rows_data) # Safe fallback

        # 2. Extract Rules
        col_rules = {} 
        sorted_rows_indices = sorted(threshold_rows_indices)
        
        for col_idx in target_columns_indices:
            rules = []
            c_name = col_names.get(col_idx, "")
            
            for r_idx in sorted_rows_indices:
                if r_idx >= len(rows_data): continue
                row = rows_data[r_idx]
                if col_idx >= len(row): continue
                
                cell = row[col_idx]
                # In rows_data, color is '#RRGGBB' or None.
                bg_color = cell['bg_color']
                if not bg_color: continue
                # We need hex without #
                hex_color = bg_color.replace('#', '')
                
                val = cell['value']
                if val is None: continue
                val_str = str(val).strip()
                
                parsed = self.parse_cell_value(val_str, col_name=c_name)
                if parsed:
                    rule = parsed
                    rule['hex'] = hex_color
                    rule['row_idx'] = r_idx 
                    rules.append(rule)
            
            rules = self.resolve_column_rules(rules)
            col_rules[col_idx] = rules
            
        # 3. Generate Preview Rows
        preview_rows = []
        
        # Slicing is fast
        start_idx = final_start_row - 1 # 1-based start row -> 0-based index
        if start_idx < 0: start_idx = 0
        
        data_slice = rows_data[start_idx:]
        if limit:
            data_slice = data_slice[:limit]
            
        consecutive_empty = 0
        MAX_EMPTY = 10
        
        for rel_idx, row in enumerate(data_slice):
            
            # Check Empty
            is_empty = True
            for c in row:
                if c['value'] is not None and str(c['value']).strip() != "":
                    is_empty = False
                    break
            if is_empty:
                consecutive_empty += 1
                if consecutive_empty >= MAX_EMPTY:
                    break
            else:
                consecutive_empty = 0

            row_out = {}
            current_abs_row_idx = start_idx + rel_idx # 0-based absolute index
            
            for i, cell in enumerate(row):
                val = cell['value']
                bg = cell['bg_color']
                if bg: bg = bg.replace('#', '')
                
                row_out[f"Col {i}"] = {'value': str(val) if val is not None else "", 'bg': None}
                
                # --- Manual Override ---
                if manual_overrides and (current_abs_row_idx, i) in manual_overrides:
                    ov = manual_overrides[(current_abs_row_idx, i)]
                    row_out[f"Col {i}"]['bg'] = ov
                    continue
                
                
                # --- Existing Non-White Color ---
                # Preserve pre-colored cells (skip threshold application)
                if bg and not self._is_white_color(bg):
                    row_out[f"Col {i}"]['bg'] = bg
                    continue
                
                # --- Apply Rules ---
                if i in target_columns_indices:
                     try:
                        v_str = str(val if val is not None else "")
                        if '<' in v_str:
                            f_val = float(v_str.replace('<','').replace('>','').replace(',','')) - 1e-9
                        else:
                            f_val = float(v_str.replace('<','').replace('>','').replace(',',''))
                        
                        matches = []
                        for rule in col_rules[i]:
                            is_range = (rule.get('type') == 'range')
                            min_ok = (f_val >= rule['min']) if is_range else (f_val > rule['min'])
                            max_ok = (f_val <= rule['max']) if is_range else (f_val < rule['max'])
                            
                            if min_ok and max_ok:
                                matches.append(rule)
                                
                        if matches:
                            def strictness_key(r):
                                score = r['min']
                                if r['min'] == float('-inf'):
                                    score = -r['max']
                                return (score, -r.get('row_idx', 9999))
                            
                            matches.sort(key=strictness_key)
                            winner = matches[-1]
                            row_out[f"Col {i}"]['bg'] = winner['hex']
                     except: pass
            
            preview_rows.append(row_out)
            
        return preview_rows, final_start_row
